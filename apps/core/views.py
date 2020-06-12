import json
from django.core import serializers
from django.forms.models import model_to_dict
from django.shortcuts import render
from django.http import HttpResponse
from django.http import JsonResponse
from django.http import Http404
from django.core import serializers
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
# from django.db.models import Count
from django.db import connection
from django.shortcuts import redirect
from django.shortcuts import get_object_or_404
from django.contrib.auth import views as admin_views
from apps.accounts.models import Product
from apps.accounts.forms import *
from django.db.models import *
from django.db.models.functions import Cast
from django.views.generic.edit import UpdateView

import datetime

from openpyxl import Workbook
from django.views.generic import TemplateView
from openpyxl.utils import get_column_letter

def landingPage(request):
    """ Vista que renderiza el landing page """
    return render(request, 'core/index.html', {})


def base(request):
    """ Vista que muestra el html base par pruebas """
    return render(request, 'base/base.html', {})


def dashboard(request):
    return render(request, 'core/dashboard.html', {})


def products_graphics(request):
    # best_selling = Product.products.annotate(total_orders=Count('get_orders'), total_mxn='total_orders'*price).order_by('-total_orders')
    best_selling = Product.products.annotate(
        total_orders=Count('get_orders'),
        total_mxn=Cast(
            Count('get_orders')*F('price'),
            FloatField())
    ).order_by('-total_orders')[0:7]
    worst_selling = Product.products.annotate(
        total_orders=Count('get_orders'),
        total_mxn=Cast(
            Count('get_orders')*F('price'),
            FloatField())
    ).order_by('total_orders')[0:4]
    # SELECT SUM(p.price)
    # FROM accounts_product p, accounts_order o
    # WHERE p.id = o.product_id
    # AND o.date_created > '2020-06-01';
    a_date = datetime.datetime.today()
    week_number = a_date.isocalendar()[1]
    week_close = Order.objects.filter(date_created__week=week_number).exclude(
        product_id__category='Servicio').aggregate(week_close=Sum('product_id__price'))
    print(week_close)
    month_number = a_date.month
    month_close = Order.objects.filter(date_created__month=month_number).exclude(
        product_id__category='Servicio').aggregate(month_close=Sum('product_id__price'))
    print(month_close)
    year_close = Order.objects.exclude(product_id__category='Servicio').aggregate(
        year_close=Sum('product_id__price'))
    january_close = Order.objects.filter(date_created__month=1).exclude(product_id__category='Servicio').aggregate(
        year_close=Sum('product_id__price'))
    february_close = Order.objects.filter(date_created__month=2).exclude(product_id__category='Servicio').aggregate(
        year_close=Sum('product_id__price'))
    march_close = Order.objects.filter(date_created__month=3).exclude(product_id__category='Servicio').aggregate(
        year_close=Sum('product_id__price'))
    april_close = Order.objects.filter(date_created__month=4).exclude(product_id__category='Servicio').aggregate(
        year_close=Sum('product_id__price'))
    may_close = Order.objects.filter(date_created__month=5).exclude(product_id__category='Servicio').aggregate(
        year_close=Sum('product_id__price'))
    june_close = Order.objects.filter(date_created__month=6).exclude(product_id__category='Servicio').aggregate(
        year_close=Sum('product_id__price'))
    context = {
        'is_popup': False,
        'has_permission': True,
        'site_url': '/',
        'best_selling': best_selling,
        'worst_selling': worst_selling,
        'week_close': week_close,
        'month_close': month_close,
        'year_close': year_close,
        'january_close': january_close,
        'february_close': february_close,
        'march_close': march_close,
        'april_close': april_close,
        'may_close': may_close,
        'june_close': june_close
    }
    return render(request, 'core/products_graphics.html', context)


def get_sales_products(request):
    """ Vista que regresa un json con las ventas del año """
    # sales = Order.objects.values('date_created').annotate(cantidad=Count('date_created'))
    # sales = Order.objects.values('date_created').annotate(cantidad=Count('date_created'))
    with connection.cursor() as cursor:
        cursor.execute("""SELECT o.date_created, Count(*) as cantidad 
                        FROM accounts_order o, accounts_product p
                        WHERE o.product_id = p.id AND p.category <> 'Servicio'
                        GROUP BY o.date_created;""")
        rows = cursor.fetchall()
    return JsonResponse(rows, safe=False)


def get_monthly_sales_products(request):
    " returns ventas por meses de cada producto"
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
            SELECT p.name || p.description as 'name',
                count(case when o.date_created >= '2020-01-01 00:00:00' AND o.date_created < '2020-02-01 00:00:00' then 1 end)as Enero,
                count(case when o.date_created >= '2020-02-01 00:00:00' AND o.date_created < '2020-03-01 00:00:00' then 1 end)as Febrero,
                count(case when o.date_created >= '2020-03-01 00:00:00' AND o.date_created < '2020-04-01 00:00:00' then 1 end)as Marzo,
                count(case when o.date_created >= '2020-04-01 00:00:00' AND o.date_created < '2020-05-01 00:00:00' then 1 end)as Abril,
                count(case when o.date_created >= '2020-05-01 00:00:00' AND o.date_created < '2020-06-01 00:00:00' then 1 end)as Mayo,
                count(case when o.date_created >= '2020-06-01 00:00:00' AND o.date_created < '2020-07-01 00:00:00' then 1 end)as Junio
                
            FROM accounts_order o, accounts_product p
            WHERE o.product_id = p.id
            AND p.category <> 'Servicio'
            GROUP BY o.product_id;
            """)
            rows = cursor.fetchall()

        return JsonResponse(rows, safe=False)
    except Exception as e:
        print(e)
        return JsonResponse({'data': None}, safe=False)


def services_graphics(request):
    # best_selling = Product.products.annotate(total_orders=Count('get_orders'), total_mxn='total_orders'*price).order_by('-total_orders')
    best_selling = Product.services.annotate(
        total_orders=Count('get_orders'),
        total_mxn=Cast(
            Count('get_orders')*F('price'),
            FloatField())
    ).order_by('-total_orders')
    january_close = Order.objects.filter(date_created__month=1, product_id__category='Servicio').aggregate(
        year_close=Sum('product_id__price'))
    february_close = Order.objects.filter(date_created__month=2, product_id__category='Servicio').aggregate(
        year_close=Sum('product_id__price'))
    march_close = Order.objects.filter(date_created__month=3, product_id__category='Servicio').aggregate(
        year_close=Sum('product_id__price'))
    april_close = Order.objects.filter(date_created__month=4, product_id__category='Servicio').aggregate(
        year_close=Sum('product_id__price'))
    may_close = Order.objects.filter(date_created__month=5, product_id__category='Servicio').aggregate(
        year_close=Sum('product_id__price'))
    june_close = Order.objects.filter(date_created__month=6, product_id__category='Servicio').aggregate(
        year_close=Sum('product_id__price'))
    # SELECT SUM(p.price)
    # FROM accounts_product p, accounts_order o
    # WHERE p.id = o.product_id
    # AND o.date_created > '2020-06-01';
    a_date = datetime.datetime.today()
    week_number = a_date.isocalendar()[1]
    week_close = Order.objects.filter(date_created__week=week_number, product_id__category='Servicio').aggregate(
        week_close=Sum('product_id__price'))
    print(week_close)
    month_number = a_date.month
    month_close = Order.objects.filter(date_created__month=month_number, product_id__category='Servicio').aggregate(
        month_close=Sum('product_id__price'))
    print(month_close)
    year_close = Order.objects.filter(product_id__category='Servicio').aggregate(
        year_close=Sum('product_id__price'))
    context = {
        'is_popup': False,
        'has_permission': True,
        'site_url': '/',
        'best_selling': best_selling,
        'week_close': week_close,
        'month_close': month_close,
        'year_close': year_close,
        'january_close': january_close,
        'february_close': february_close,
        'march_close': march_close,
        'april_close': april_close,
        'may_close': may_close,
        'june_close': june_close
    }
    return render(request, 'core/services_graphics.html', context)


def get_monthly_sales_services(request):
    " returns ventas por meses de cada producto"
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
            SELECT p.name,
                count(case when o.date_created >= '2020-01-01 00:00:00' AND o.date_created < '2020-02-01 00:00:00' then 1 end)as Enero,
                count(case when o.date_created >= '2020-02-01 00:00:00' AND o.date_created < '2020-03-01 00:00:00' then 1 end)as Febrero,
                count(case when o.date_created >= '2020-03-01 00:00:00' AND o.date_created < '2020-04-01 00:00:00' then 1 end)as Marzo,
                count(case when o.date_created >= '2020-04-01 00:00:00' AND o.date_created < '2020-05-01 00:00:00' then 1 end)as Abril,
                count(case when o.date_created >= '2020-05-01 00:00:00' AND o.date_created < '2020-06-01 00:00:00' then 1 end)as Mayo,
                count(case when o.date_created >= '2020-06-01 00:00:00' AND o.date_created < '2020-07-01 00:00:00' then 1 end)as Junio
                
            FROM accounts_order o, accounts_product p
            WHERE o.product_id = p.id
            AND p.category == 'Servicio'
            GROUP BY o.product_id;
            """)
            rows = cursor.fetchall()

        return JsonResponse(rows, safe=False)
    except Exception as e:
        print(e)
        return JsonResponse({'data': None}, safe=False)


def get_sales_services(request):
    """ Vista que regresa un json con las ventas del año """
    # sales = Order.objects.values('date_created').annotate(cantidad=Count('date_created'))
    # sales = Order.objects.values('date_created').annotate(cantidad=Count('date_created'))
    with connection.cursor() as cursor:
        cursor.execute("""SELECT o.date_created, Count(*) as cantidad 
                        FROM accounts_order o, accounts_product p
                        WHERE o.product_id = p.id AND p.category = 'Servicio'
                        GROUP BY o.date_created;""")
        rows = cursor.fetchall()
    return JsonResponse(rows, safe=False)


def products(request):
    products = Product.objects.all()
    return render(request, 'core/products.html', {'products': products})


def add_product(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('add_product')
        return render(request, 'core/add_product.html', {'form': form})
    form = ProductForm()
    return render(request, 'core/add_product.html', {'form': form})


def sales(request):
    sales = Order.objects.order_by('date_created')
    return render(request, 'core/sales.html', {'sales': sales})


def clients(request):
    Customer.objects.all()
    return render(request, 'core/clients.html', {})


def config(request):
    return render(request, 'core/config.html', {})


class UpdateCustomer(UpdateView):
    model = Customer
    fields = '__all__'
    template_name_suffix = '_update_form'
    success_url = None

from openpyxl.styles import Font, Alignment, Border, Side, NamedStyle, Fill, PatternFill

class SalesReport(TemplateView):
    FORMAT_CURRENCY = "_($* #,##0.00_);_($* (#,##0.00);_($* ""-""??_);_(@_)"
    font_title = Font(size = 14, name = 'Poppins', color="666666")
    font_subtitle = Font(size = 11, name = 'Roboto', color="666666")
    v_alignment_center = Alignment(vertical="center")
    alignment_center = Alignment(horizontal="center", vertical="center")
    t_border = Border(top=Side(style='thin', color='DEE2E6'))
    base_fill = PatternFill("solid", fgColor="F9F9F9")
    header_style = NamedStyle(name="header_style", font=font_subtitle, border = t_border, alignment = v_alignment_center)
    data_style = NamedStyle(name="data_style", border = t_border, alignment = v_alignment_center, fill=base_fill)
    data_style_no_bg = NamedStyle(name="data_style_no_bg", border = t_border, alignment = v_alignment_center)
    def get(self, request, *args, **kwargs):
        a_date = datetime.datetime.today()
        week_number = a_date.isocalendar()[1]
        week_close = Order.objects.filter(date_created__week=week_number).exclude(
            product_id__category='Servicio').aggregate(week_close=Sum('product_id__price'))
        month_number = a_date.month
        month_close = Order.objects.filter(date_created__month=month_number).exclude(
            product_id__category='Servicio').aggregate(month_close=Sum('product_id__price'))
        year_close = Order.objects.exclude(product_id__category='Servicio').aggregate(
            year_close=Sum('product_id__price'))

        s_week_close = Order.objects.filter(date_created__week=week_number, product_id__category='Servicio').aggregate(
            week_close=Sum('product_id__price'))
        s_month_close = Order.objects.filter(date_created__month=month_number, product_id__category='Servicio').aggregate(
            month_close=Sum('product_id__price'))
        s_year_close = Order.objects.filter(product_id__category='Servicio').aggregate(
            year_close=Sum('product_id__price'))
        # Close for products
        p_january_close = Order.objects.filter(date_created__month=1).exclude(product_id__category='Servicio').aggregate(
            year_close=Sum('product_id__price'))
        p_february_close = Order.objects.filter(date_created__month=2).exclude(product_id__category='Servicio').aggregate(
            year_close=Sum('product_id__price'))
        p_march_close = Order.objects.filter(date_created__month=3).exclude(product_id__category='Servicio').aggregate(
            year_close=Sum('product_id__price'))
        p_april_close = Order.objects.filter(date_created__month=4).exclude(product_id__category='Servicio').aggregate(
            year_close=Sum('product_id__price'))
        p_may_close = Order.objects.filter(date_created__month=5).exclude(product_id__category='Servicio').aggregate(
            year_close=Sum('product_id__price'))
        p_june_close = Order.objects.filter(date_created__month=6).exclude(product_id__category='Servicio').aggregate(
            year_close=Sum('product_id__price'))

        # Close for services
        january_close = Order.objects.filter(date_created__month=1, product_id__category='Servicio').aggregate(
            year_close=Sum('product_id__price'))
        february_close = Order.objects.filter(date_created__month=2, product_id__category='Servicio').aggregate(
            year_close=Sum('product_id__price'))
        march_close = Order.objects.filter(date_created__month=3, product_id__category='Servicio').aggregate(
            year_close=Sum('product_id__price'))
        april_close = Order.objects.filter(date_created__month=4, product_id__category='Servicio').aggregate(
            year_close=Sum('product_id__price'))
        may_close = Order.objects.filter(date_created__month=5, product_id__category='Servicio').aggregate(
            year_close=Sum('product_id__price'))
        june_close = Order.objects.filter(date_created__month=6, product_id__category='Servicio').aggregate(
            year_close=Sum('product_id__price'))

        # Best selling for porducts
        best_selling = Product.products.annotate(
            total_orders=Count('get_orders'),
            total_mxn=Cast(
                Count('get_orders')*F('price'),
                FloatField())
        ).order_by('-total_orders')[0:7]

        worst_selling = Product.products.annotate(
            total_orders=Count('get_orders'),
            total_mxn=Cast(
                Count('get_orders')*F('price'),
                FloatField())
        ).order_by('total_orders')[0:4]

        services_sales = Product.services.annotate(
            total_orders=Count('get_orders'),
            total_mxn=Cast(
                Count('get_orders')*F('price'),
                FloatField())
        ).order_by('-total_orders')
        # create book
        wb = Workbook()
        ws = wb.active
        ws.sheet_view.showGridLines = False
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        ws['B2'] = 'REPORTE ANUAL ENERO -JUNIO 2019'
        ws['B2'].font = self.font_title
        ws['B2'].alignment = self.alignment_center
        ws.merge_cells('B2:H2')

        ws['B4'] = 'CONSULTA DE LAS VENTAS (PRODUCTO)'
        ws['B4'].font = self.font_title
        ws['B4'].alignment = self.alignment_center
        ws.merge_cells('B4:D4')

        ws['F4'] = 'CONSULTA DE LAS VENTAS (SERVICIO)'
        ws['F4'].font = self.font_title
        ws['F4'].alignment = self.alignment_center
        ws.merge_cells('F4:H4')

        ws['B6'] = 'CIERRE SEMANAL'
        ws['B6'].style = self.header_style
        ws['C6'] = 'CIERRE MENSUAL'
        ws['C6'].style = self.header_style
        ws['D6'] = 'CIERRE DEL AÑO'
        ws['D6'].style = self.header_style

        ws['B7'] = week_close['week_close']
        ws['B7'].style = self.data_style
        ws['B7'].number_format = self.FORMAT_CURRENCY
        ws['C7'] = month_close['month_close']
        ws['C7'].style = self.data_style
        ws['C7'].number_format = self.FORMAT_CURRENCY
        ws['D7'] = year_close['year_close']
        ws['D7'].style = self.data_style
        ws['D7'].number_format = self.FORMAT_CURRENCY

        ws['F6'] = 'CIERRE SEMANAL'
        ws['F6'].style = self.header_style
        ws['G6'] = 'CIERRE MENSUAL'
        ws['G6'].style = self.header_style
        ws['H6'] = 'CIERRE DEL AÑO'
        ws['H6'].style = self.header_style

        ws['F7'] = s_week_close['week_close']
        ws['F7'].style = self.data_style
        ws['F7'].number_format = self.FORMAT_CURRENCY
        ws['G7'] = s_month_close['month_close']
        ws['G7'].style = self.data_style
        ws['G7'].number_format = self.FORMAT_CURRENCY
        ws['H7'] = s_year_close['year_close']
        ws['H7'].style = self.data_style
        ws['H7'].number_format = self.FORMAT_CURRENCY

        # month close for porducts
        ws['B9'] = 'MES'
        ws['B9'].style = self.header_style
        ws['C9'] = 'CIERRE'
        ws['C9'].style = self.header_style
        ws['D10'] = 'INCREMENTO'
        ws['D10'].style = self.header_style
        ws['B10'] = 'Enero'
        ws['B10'].style = self.data_style
        ws['C10'] = p_january_close['year_close']
        ws['C10'].style = self.data_style
        ws['C10'].number_format = self.FORMAT_CURRENCY
        ws['B11'] = 'Febrero'
        ws['B11'].style = self.data_style_no_bg
        ws['C11'] = p_february_close['year_close']
        ws['C11'].style = self.data_style_no_bg
        ws['C11'].number_format = self.FORMAT_CURRENCY
        ws['D11'] = '=((C11*100/C10)-100)/100'
        ws['D11'].style = self.data_style_no_bg
        ws['D11'].number_format = '0.00%'
        ws['B12'] = 'Marzo'
        ws['B12'].style = self.data_style
        ws['C12'] = p_march_close['year_close']
        ws['C12'].style = self.data_style
        ws['C12'].number_format = self.FORMAT_CURRENCY
        ws['D12'] = '=((C12*100/C11)-100)/100'
        ws['D12'].style = self.data_style
        ws['D12'].number_format = '0.00%'
        ws['B13'] = 'Abril'
        ws['B13'].style = self.data_style_no_bg
        ws['C13'] = p_april_close['year_close']
        ws['C13'].style = self.data_style_no_bg
        ws['C13'].number_format = self.FORMAT_CURRENCY
        ws['D13'] = '=((C13*100/C12)-100)/100'
        ws['D13'].style = self.data_style_no_bg
        ws['D13'].number_format = '0.00%'
        ws['B14'] = 'Mayo'
        ws['B14'].style = self.data_style
        ws['C14'] = p_may_close['year_close']
        ws['C14'].style = self.data_style
        ws['C14'].number_format = self.FORMAT_CURRENCY
        ws['D14'] = '=((C14*100/C13)-100)/100'
        ws['D14'].style = self.data_style
        ws['D14'].number_format = '0.00%'
        ws['B15'] = 'Junio'
        ws['B15'].style = self.data_style_no_bg
        ws['C15'] = p_june_close['year_close']
        ws['C15'].style = self.data_style_no_bg
        ws['C15'].number_format = self.FORMAT_CURRENCY
        ws['D15']='=((C15*100/C14)-100)/100'
        ws['D15'].style = self.data_style_no_bg
        ws['D15'].number_format = '0.00%'
        # TOTAL
        ws['B16'] = 'Total:'
        ws['B16'].style = self.data_style
        ws['C16'] = year_close['year_close']
        ws['C16'].style = self.data_style
        ws['C16'].number_format = self.FORMAT_CURRENCY
        # month close for services
        ws['F9'] = 'MES'
        ws['F9'].style = self.header_style
        ws['G9'] = 'CIERRE'
        ws['G9'].style = self.header_style
        ws['H10'] = 'INCREMENTO'
        ws['H10'].style = self.header_style
        ws['F10'] = 'Enero'
        ws['F10'].style = self.data_style
        ws['G10'] = january_close['year_close']
        ws['G10'].style = self.data_style
        ws['G10'].number_format = self.FORMAT_CURRENCY
        ws['F11'] = 'Febrero'
        ws['F11'].style = self.data_style_no_bg
        ws['G11'] = february_close['year_close']
        ws['G11'].style = self.data_style_no_bg
        ws['G11'].number_format = self.FORMAT_CURRENCY

        ws['H11']='=((G11*100/G10)-100)/100'
        ws['H11'].style = self.data_style_no_bg
        ws['H11'].number_format = '0.00%'

        ws['F12'] = 'Marzo'
        ws['F12'].style = self.data_style
        ws['G12'] = march_close['year_close']
        ws['G12'].style = self.data_style
        ws['G12'].number_format = self.FORMAT_CURRENCY

        ws['H12']='=((G12*100/G11)-100)/100'
        ws['H12'].style = self.data_style
        ws['H12'].number_format = '0.00%'

        ws['F13'] = 'Abril'
        ws['F13'].style = self.data_style_no_bg
        ws['G13'] = april_close['year_close']
        ws['G13'].style = self.data_style_no_bg
        ws['G13'].number_format = self.FORMAT_CURRENCY

        ws['H13']='=((G13*100/G12)-100)/100'
        ws['H13'].style = self.data_style_no_bg
        ws['H13'].number_format = '0.00%'

        ws['F14'] = 'Mayo'
        ws['F14'].style = self.data_style
        ws['G14'] = may_close['year_close']
        ws['G14'].style = self.data_style
        ws['G14'].number_format = self.FORMAT_CURRENCY

        ws['H14']='=((G14*100/G13)-100)/100'
        ws['H14'].style = self.data_style
        ws['H14'].number_format = '0.00%'

        ws['F15'] = 'Junio'
        ws['F15'].style = self.data_style_no_bg
        ws['G15'] = june_close['year_close']
        ws['G15'].style = self.data_style_no_bg
        ws['G15'].number_format = self.FORMAT_CURRENCY

        ws['H15']='=((G15*100/G14)-100)/100'
        ws['H15'].style = self.data_style
        ws['H15'].number_format = '0.00%'

        # TOTAL
        ws['F16'] = 'Total:'
        ws['F16'].style = self.data_style
        ws['G16'] = s_year_close['year_close']
        ws['G16'].style = self.data_style
        ws['G16'].number_format = self.FORMAT_CURRENCY

        ws['B19'] = 'CONSULTA DE PRODUCTOS MÁS DEMANDADOS'
        ws['B19'].font = self.font_title
        ws['B19'].alignment = self.alignment_center
        ws.merge_cells(start_row=19, start_column=2, end_row=19, end_column=7)
        ws['B20'] = 'Producto'
        ws['B20'].style = self.header_style
        ws['C20'] = 'Descripción'
        ws['C20'].style = self.header_style
        ws['D20'] = 'Categoría'
        ws['D20'].style = self.header_style
        ws['E20'] = 'Precio'
        ws['E20'].style = self.header_style
        ws['F20'] = 'Total ventas'
        ws['F20'].style = self.header_style
        ws['G20'] = 'Total MXN'
        ws['G20'].style = self.header_style
        count = 21
        flag = True
        for sell in best_selling:
            if flag:
                ws.cell(row=count, column=2).value = sell.name
                ws.cell(row=count, column=2).style = self.data_style
                ws.cell(row=count, column=3).value = sell.description
                ws.cell(row=count, column=3).style = self.data_style
                ws.cell(row=count, column=4).value = sell.category
                ws.cell(row=count, column=4).style = self.data_style
                ws.cell(row=count, column=5).value = sell.price
                ws.cell(row=count, column=5).style = self.data_style
                ws.cell(row=count, column=5).number_format = self.FORMAT_CURRENCY
                ws.cell(row=count, column=6).value = sell.total_orders
                ws.cell(row=count, column=6).style = self.data_style
                ws.cell(row=count, column=7).value = sell.total_mxn
                ws.cell(row=count, column=7).style = self.data_style
                ws.cell(row=count, column=7).number_format = self.FORMAT_CURRENCY
            else:
                ws.cell(row=count, column=2).value = sell.name
                ws.cell(row=count, column=2).style = self.data_style_no_bg
                ws.cell(row=count, column=3).value = sell.description
                ws.cell(row=count, column=3).style = self.data_style_no_bg
                ws.cell(row=count, column=4).value = sell.category
                ws.cell(row=count, column=4).style = self.data_style_no_bg
                ws.cell(row=count, column=5).value = sell.price
                ws.cell(row=count, column=5).style = self.data_style_no_bg
                ws.cell(row=count, column=5).number_format = self.FORMAT_CURRENCY
                ws.cell(row=count, column=6).value = sell.total_orders
                ws.cell(row=count, column=6).style = self.data_style_no_bg
                ws.cell(row=count, column=7).value = sell.total_mxn
                ws.cell(row=count, column=7).style = self.data_style_no_bg
                ws.cell(row=count, column=7).number_format = self.FORMAT_CURRENCY
            flag = not(flag)
            count += 1

        count += 1
        ws.cell(row=count, column=2).value = "CONSULTA DE PRODUCTOS MENOS DEMANDADOS"
        ws.cell(row=count, column=2).font = self.font_title
        ws.cell(row=count, column=2).alignment = self.alignment_center
        ws.merge_cells(start_row=count, start_column=2,
                       end_row=count, end_column=7)
        count += 1
        ws.cell(row=count, column=2).value = 'Producto'
        ws.cell(row=count, column=2).style = self.header_style
        ws.cell(row=count, column=3).value = 'Descripción'
        ws.cell(row=count, column=3).style = self.header_style
        ws.cell(row=count, column=4).value = 'Categoría'
        ws.cell(row=count, column=4).style = self.header_style
        ws.cell(row=count, column=5).value = 'Precio'
        ws.cell(row=count, column=5).style = self.header_style
        ws.cell(row=count, column=6).value = 'Total ventas'
        ws.cell(row=count, column=6).style = self.header_style
        ws.cell(row=count, column=7).value = 'Total MXN'
        ws.cell(row=count, column=7).style = self.header_style
        count += 1
        flag = True
        for sell in worst_selling:
            if flag:
                ws.cell(row=count, column=2).value = sell.name
                ws.cell(row=count, column=2).style = self.data_style
                ws.cell(row=count, column=3).value = sell.description
                ws.cell(row=count, column=3).style = self.data_style
                ws.cell(row=count, column=4).value = sell.category
                ws.cell(row=count, column=4).style = self.data_style
                ws.cell(row=count, column=5).value = sell.price
                ws.cell(row=count, column=5).style = self.data_style
                ws.cell(row=count, column=5).number_format = self.FORMAT_CURRENCY
                ws.cell(row=count, column=6).value = sell.total_orders
                ws.cell(row=count, column=6).style = self.data_style
                ws.cell(row=count, column=7).value = sell.total_mxn
                ws.cell(row=count, column=7).style = self.data_style
                ws.cell(row=count, column=7).number_format = self.FORMAT_CURRENCY
            else:
                ws.cell(row=count, column=2).value = sell.name
                ws.cell(row=count, column=2).style = self.data_style_no_bg
                ws.cell(row=count, column=3).value = sell.description
                ws.cell(row=count, column=3).style = self.data_style_no_bg
                ws.cell(row=count, column=4).value = sell.category
                ws.cell(row=count, column=4).style = self.data_style_no_bg
                ws.cell(row=count, column=5).value = sell.price
                ws.cell(row=count, column=5).style = self.data_style_no_bg
                ws.cell(row=count, column=5).number_format = self.FORMAT_CURRENCY
                ws.cell(row=count, column=6).value = sell.total_orders
                ws.cell(row=count, column=6).style = self.data_style_no_bg
                ws.cell(row=count, column=7).value = sell.total_mxn
                ws.cell(row=count, column=7).style = self.data_style_no_bg
                ws.cell(row=count, column=7).number_format = self.FORMAT_CURRENCY
            flag = not(flag)
            count += 1

        count += 1
        ws.cell(row=count, column=2).value = "SERVICIOS CONTRATADOS"
        ws.cell(row=count, column=2).alignment = self.alignment_center
        ws.cell(row=count, column=2).font = self.font_title
        ws.merge_cells(start_row=count, start_column=2,
                       end_row=count, end_column=7)
        count += 1
        ws.cell(row=count, column=2).value = 'Producto'
        ws.cell(row=count, column=2).style = self.header_style
        ws.cell(row=count, column=3).value = 'Categoría'
        ws.cell(row=count, column=3).style = self.header_style
        ws.cell(row=count, column=4).value = 'Precio'
        ws.cell(row=count, column=4).style = self.header_style
        ws.cell(row=count, column=5).value = 'Total ventas'
        ws.cell(row=count, column=5).style = self.header_style
        ws.cell(row=count, column=6).value = 'Total MXN'
        ws.cell(row=count, column=6).style = self.header_style
        count += 1
        flag = True
        for sell in services_sales:
            if flag:
                ws.cell(row=count, column=2).value = sell.name
                ws.cell(row=count, column=2).style = self.data_style
                ws.cell(row=count, column=3).value = sell.category
                ws.cell(row=count, column=3).style = self.data_style
                ws.cell(row=count, column=4).value = sell.price
                ws.cell(row=count, column=4).style = self.data_style
                ws.cell(row=count, column=4).number_format = self.FORMAT_CURRENCY
                ws.cell(row=count, column=5).value = sell.total_orders
                ws.cell(row=count, column=5).style = self.data_style
                ws.cell(row=count, column=6).value = sell.total_mxn
                ws.cell(row=count, column=6).style = self.data_style
                ws.cell(row=count, column=6).number_format = self.FORMAT_CURRENCY
            else:
                ws.cell(row=count, column=2).value = sell.name
                ws.cell(row=count, column=2).style = self.data_style_no_bg
                ws.cell(row=count, column=3).value = sell.category
                ws.cell(row=count, column=3).style = self.data_style_no_bg
                ws.cell(row=count, column=4).value = sell.price
                ws.cell(row=count, column=4).style = self.data_style_no_bg
                ws.cell(row=count, column=4).number_format = self.FORMAT_CURRENCY
                ws.cell(row=count, column=5).value = sell.total_orders
                ws.cell(row=count, column=5).style = self.data_style_no_bg
                ws.cell(row=count, column=6).value = sell.total_mxn
                ws.cell(row=count, column=6).style = self.data_style_no_bg
                ws.cell(row=count, column=6).number_format = self.FORMAT_CURRENCY
            flag = not(flag)
            count += 1
        
        
        for i in range(0,ws.max_row):
            ws.row_dimensions[i+1].height=25
        # return file
        file_name = "Reporte_anual.xlsx"
        response = HttpResponse(content_type='application/ms-excel')
        content = "attachment; filename = {0}".format(file_name)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
